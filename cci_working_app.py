import streamlit as st
import pandas as pd
import numpy as np
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime
from PIL import Image
import os
import io

# OpenPyXL formatting styling ke liye
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# FPDF2 Integration
from fpdf import FPDF

# --- PAGE SETUP & THEME ---
st.set_page_config(page_title="CCI Calculation Working Utility", layout="wide")

# --- 🔒 SIMPLE PASSWORD PROTECTION SYSTEM ---
def check_password():
    """Returns True if the user had the correct password."""
    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if st.session_state["utility_access_key"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["utility_access_key"]  # cache clear karne ke liye
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # Login screen UI
        st.markdown("<h2 style='text-align: center; color: #8A2BE2;'>🔒 Softview CCI Utility Login</h2>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            # autocomplete="new-password" se google suggestion prompt aana band ho jayega
            st.text_input(
                "System Access Key", 
                type="password", 
                on_change=password_entered, 
                key="utility_access_key",
                autocomplete="new-password"
            )
            if "password_correct" in st.session_state and not st.session_state["password_correct"]:
                st.error("❌ Invalid Key. Please try again.")
        return False
    elif not st.session_state["password_correct"]:
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.text_input(
                "System Access Key", 
                type="password", 
                on_change=password_entered, 
                key="utility_access_key",
                autocomplete="new-password"
            )
            st.error("❌ Invalid Key. Please try again.")
        return False
    else:
        return True
# Agar password check fail ho jaye, toh aage ka code mat chalao (Stop right here)
if not check_password():
    st.stop()

# --- 🚀 AB YAHAAN SE AAPKA BAAKI KA PURA CODE SHURU HOGA ---
st.markdown("""
    <style>
    /* Ye code deploy menu (GitHub icon) ko hide kar dega */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)

st.markdown("""
<style>

/* =========================
   MAIN APP BACKGROUND
========================= */

.stApp {
    background: linear-gradient(
        135deg,
        #D6D6D6 0%,
        #C9C9C9 50%,
        #BEBEBE 100%
    );
    font-family: 'Segoe UI', sans-serif;
}

/* =========================
   HEADER HIDE
========================= */

header[data-testid="stHeader"] {
    display: none;
}

footer {
    display: none;
}

/* =========================
   SIDEBAR ULTRA PREMIUM
========================= */

[data-testid="stSidebar"] {
    background: linear-gradient(
        180deg,
        #2B2B2B 0%,
        #3A3A3A 100%
    ) !important;

    border-right: 2px solid #FF1493;

    box-shadow:
        8px 0px 30px rgba(255, 20, 147, 0.35);

    padding-top: 10px;
}

/* Sidebar Text */

[data-testid="stSidebar"] * {
    color: white !important;
}

/* Sidebar Header */

.sidebar-header {
    background: linear-gradient(
        135deg,
        #FF1493,
        #C71585
    );

    padding: 14px;
    border-radius: 14px;

    text-align: center;

    font-size: 20px;
    font-weight: 700;

    margin-bottom: 20px;

    box-shadow:
        0 0 25px rgba(255,20,147,0.5);
}

/* =========================
   MAIN TITLE
========================= */

h1 {
    color: #FF1493 !important;

    text-align: center;

    font-size: 42px !important;

    font-weight: 900 !important;

    text-shadow:
        0px 0px 15px rgba(255,20,147,0.5);
}

/* =========================
   SUBHEADINGS
========================= */

h2, h3, h4 {
    color: #C71585 !important;
    font-weight: 800 !important;
}

/* =========================
   INPUT FIELDS
========================= */

.stTextInput input,
.stNumberInput input,
.stDateInput input {

    background-color: #ECECEC !important;

    border: 2px solid #FF1493 !important;

    border-radius: 14px !important;

    color: #2B2B2B !important;

    font-weight: 600 !important;

    box-shadow:
        0px 0px 10px rgba(255,20,147,0.15);
}

/* =========================
   BUTTONS
========================= */

.stButton > button,
.stDownloadButton > button {

    background: linear-gradient(
        135deg,
        #FF1493,
        #C71585
    ) !important;

    color: white !important;

    border: none !important;

    border-radius: 16px !important;

    font-size: 16px !important;

    font-weight: 700 !important;

    padding: 12px 20px !important;

    transition: all 0.3s ease !important;

    box-shadow:
        0px 8px 25px rgba(255,20,147,0.35);
}

.stButton > button:hover,
.stDownloadButton > button:hover {

    transform: translateY(-3px);

    box-shadow:
        0px 12px 30px rgba(255,20,147,0.55);

    background: linear-gradient(
        135deg,
        #FF3CAC,
        #784BA0
    ) !important;
}

/* =========================
   TABS ULTRA PREMIUM
========================= */

.stTabs [data-baseweb="tab-list"] {

    gap: 12px;

    background: rgba(255,255,255,0.15);

    padding: 10px;

    border-radius: 18px;

    box-shadow:
        inset 0px 0px 15px rgba(255,255,255,0.1);
}

/* Inactive Tabs */

.stTabs [data-baseweb="tab"] {

    background: #4A4A4A !important;

    color: white !important;

    border-radius: 16px !important;

    padding: 14px 24px !important;

    font-weight: 700 !important;

    transition: 0.3s ease;

    border: 2px solid transparent;

    box-shadow:
        0px 5px 15px rgba(0,0,0,0.2);
}

/* Active Tab */

.stTabs [aria-selected="true"] {

    background: linear-gradient(
        135deg,
        #FF1493,
        #C71585
    ) !important;

    color: white !important;

    border: 2px solid #FFB6D9 !important;

    box-shadow:
        0px 0px 25px rgba(255,20,147,0.6);

    transform: translateY(-2px);
}

/* =========================
   EXPANDERS PREMIUM
========================= */

.streamlit-expanderHeader {

    background: linear-gradient(
        135deg,
        #FF1493,
        #C71585
    ) !important;

    color: white !important;

    border-radius: 14px !important;

    font-size: 16px !important;

    font-weight: 700 !important;

    border: none !important;

    padding: 10px !important;

    box-shadow:
        0px 5px 20px rgba(255,20,147,0.4);
}

.streamlit-expanderContent {

    background: rgba(255,255,255,0.85);

    border-radius: 0px 0px 16px 16px;

    padding: 15px;

    border: 1px solid rgba(255,20,147,0.2);
}

/* =========================
   DATAFRAMES
========================= */

[data-testid="stDataFrame"] {

    border-radius: 18px !important;

    overflow: hidden;

    box-shadow:
        0px 10px 30px rgba(0,0,0,0.15);
}

/* =========================
   METRIC CARDS
========================= */

.luxury-card {

    background: rgba(255,255,255,0.92);

    border-left: 8px solid #FF1493;

    border-radius: 20px;

    padding: 25px;

    backdrop-filter: blur(10px);

    box-shadow:
        0px 10px 30px rgba(255,20,147,0.18);

    transition: all 0.3s ease;
}

.luxury-card:hover {

    transform: translateY(-5px);

    box-shadow:
        0px 15px 40px rgba(255,20,147,0.35);
}

.luxury-metric-title {

    color: #666666;

    font-size: 14px;

    font-weight: 700;

    letter-spacing: 1px;
}

.luxury-metric-value {

    color: #C71585;

    font-size: 30px;

    font-weight: 900;

    margin-top: 8px;
}

/* =========================
   TABLES
========================= */

.invoice-table {

    width: 100%;

    border-collapse: collapse;

    overflow: hidden;

    border-radius: 18px;
}

.invoice-table th {

    background: linear-gradient(
        135deg,
        #FF1493,
        #C71585
    );

    color: white;

    padding: 14px;

    font-size: 15px;
}

.invoice-table td {

    padding: 12px;

    background: rgba(255,255,255,0.92);

    border-bottom: 1px solid #E0E0E0;
}

/* =========================
   SCROLLBAR
========================= */

::-webkit-scrollbar {
    width: 10px;
}

::-webkit-scrollbar-track {
    background: #D3D3D3;
}

::-webkit-scrollbar-thumb {

    background: linear-gradient(
        #FF1493,
        #C71585
    );

    border-radius: 10px;
}
/* =========================
   SIDEBAR FORM BUTTON FIX
========================= */

section[data-testid="stSidebar"] .stButton > button,
section[data-testid="stSidebar"] .stDownloadButton > button,
section[data-testid="stSidebar"] button[kind="primary"] {

    background: linear-gradient(
        135deg,
        #FF1493,
        #C71585
    ) !important;

    color: white !important;

    border: none !important;

    border-radius: 16px !important;

    min-height: 50px !important;

    font-size: 16px !important;

    font-weight: 800 !important;

    width: 100% !important;

    margin-top: 15px !important;

    box-shadow:
        0px 8px 24px rgba(255,20,147,0.45) !important;

    transition: all 0.3s ease !important;
}

section[data-testid="stSidebar"] .stButton > button:hover,
section[data-testid="stSidebar"] button[kind="primary"]:hover {

    transform: translateY(-3px);

    background: linear-gradient(
        135deg,
        #FF3CAC,
        #C71585
    ) !important;

    box-shadow:
        0px 12px 28px rgba(255,20,147,0.65) !important;
}

section[data-testid="stSidebar"] .stButton > button:active {

    transform: scale(0.98);
}
/* =========================
   EXPANDED SECTION FIX
========================= */

/* Expanded Area Background */

details[open] > div {

    background: rgba(255,255,255,0.96) !important;

    border-radius: 0px 0px 16px 16px !important;

    padding: 14px !important;

    color: #2B2B2B !important;
}

/* ALL TEXT INSIDE EXPANDED AREA */

details[open] label,
details[open] p,
details[open] span,
details[open] div {

    color: #2B2B2B !important;

    font-weight: 600 !important;
}

/* Input Labels */

details[open] .stTextInput label,
details[open] .stNumberInput label,
details[open] .stDateInput label {

    color: #C71585 !important;

    font-weight: 800 !important;
}

/* Save Button Fix */

details[open] button {

    background: linear-gradient(
        135deg,
        #FF1493,
        #C71585
    ) !important;

    color: white !important;

    border-radius: 14px !important;

    border: none !important;

    font-weight: 800 !important;

    min-height: 48px !important;

    box-shadow:
        0px 8px 20px rgba(255,20,147,0.35);
}
/* =====================================
   FORCE SIDEBAR SAVE BUTTON FIX
===================================== */

details[open] div[data-testid="stForm"] button {

    background: linear-gradient(
        135deg,
        #FF1493 0%,
        #C71585 100%
    ) !important;

    color: white !important;

    border: none !important;

    border-radius: 16px !important;

    height: 52px !important;

    width: 100% !important;

    font-size: 16px !important;

    font-weight: 800 !important;

    margin-top: 18px !important;

    box-shadow:
        0px 8px 22px rgba(255,20,147,0.45) !important;

    transition: all 0.3s ease !important;

    opacity: 1 !important;
}

/* Hover */

details[open] div[data-testid="stForm"] button:hover {

    background: linear-gradient(
        135deg,
        #FF3CAC,
        #C71585
    ) !important;

    transform: translateY(-2px);

    box-shadow:
        0px 12px 28px rgba(255,20,147,0.65) !important;
}

/* Button Text */

details[open] div[data-testid="stForm"] button p,
details[open] div[data-testid="stForm"] button span {

    color: white !important;

    font-weight: 800 !important;

    opacity: 1 !important;
}
/* =========================================
   FINAL FORCE SAVE BUTTON FIX
========================================= */

/* Submit Button Exact Target */

div[data-testid="stFormSubmitButton"] > button {

    background: linear-gradient(
        135deg,
        #FF1493 0%,
        #C71585 100%
    ) !important;

    color: white !important;

    border: none !important;

    border-radius: 18px !important;

    min-height: 52px !important;

    width: 100% !important;

    font-size: 16px !important;

    font-weight: 900 !important;

    margin-top: 20px !important;

    opacity: 1 !important;

    box-shadow:
        0px 10px 26px rgba(255,20,147,0.45) !important;

    transition: all 0.3s ease !important;
}

/* Text inside button */

div[data-testid="stFormSubmitButton"] > button p,
div[data-testid="stFormSubmitButton"] > button span {

    color: white !important;

    font-weight: 900 !important;
}

/* Hover */

div[data-testid="stFormSubmitButton"] > button:hover {

    background: linear-gradient(
        135deg,
        #FF3CAC,
        #D414A0
    ) !important;

    transform: translateY(-2px);

    box-shadow:
        0px 14px 30px rgba(255,20,147,0.65) !important;
}

/* Active */

div[data-testid="stFormSubmitButton"] > button:active {

    transform: scale(0.98);
}

</style>
""", unsafe_allow_html=True)

# --- DATABASE INITIALIZATION ---
def init_db():
    if not firebase_admin._apps:
        try:
            firebase_info = {
                "type": "service_account",
                "project_id": "softview-cci-utility",
                "private_key_id": "2de8f75490ad61288e60176a897a0d2552d9f6a2",
                "private_key": """-----BEGIN PRIVATE KEY-----
MIIEvQIBADANBgkqhkiG9w0BAQEFAASCBKcwggSjAgEAAoIBAQDJFOsTRzYI9UHY
sQELAiIQ+oSE5InQu7Qid8TcDq+S/9qgo/JYrOpfUhHifpo7237frnzgLfwbRg3B
HYQF9KrQRNgyo3ldJ/37/SVnh68EsTubQTUAj0+OoIU4IdDi6DX6UjgY87lVh/It
3dkCRQgr6R5VpRZAPe3o7sBKiUyW+OQRYH+ivB/wL6rasMsZ4N3Uia4Q/a7FYjMr
J8lF0EPgK3G7mBo41OH6JlCUF6+JdOarLmTePE0I7yV2JaRLprZHcuHOt0FQ6mAt
1i7Fykwt8AyulqhosBXx++G0M/4m/FlZ1hLHyg/GZuXQCM4aso3jiY++64RM/G3W
/JMgXIm/AgMBAAECggEARZ6O90k//aEAr8g07r1J0ywRZruDtbfPWeGy0GfrRQ99
gAHTXf1cVi1hkNQ5jUDlMAfHJ4z0TPMXvwV7mN5Jp7s2SW4NDoJSrwTKBwrUv2Cn
nQzaD3wO1Phht3oJyw4P5j/COW1k/SaI5HV7dYw2FJHtVUTHq+1lKjFDdWsND/W1
QFLMVPlL9GiRzw+gnVMu4lbXKTD1KbasjPRu1rafnwUp2+1/7Rl8uNP8EDasBq67
v7pHSZpRJ+gAKRy7p56DSHZRSq709PDU+ht37yPhhXU3EPt/FFTg6i94vVDOsQIk
OlBlb4XpPPYjc2B5jZXxqy43dozRtEOlLST9jCtjoQKBgQD+5e7P/Edmwy1H3WWw
O1itMNnPLk+CAbndPTn8kxTubGK6xY6MK4HgEO88TxhUm0fQlMfAndlx6pd0Xrtk
+m9mQyuVAb4ILZ722SNh9IiWtVGoq9UMHjhtsPoOfCQGBnErVbYMysYBfqPIdn54
ctnoyI5tDWXuBZGCYx2RiYpm4QKBgQDJ827KXRA9HpWuAPAyOs5ixZfhAZTE1Imj
qaYdTugkpMS0iMxpkGnrCFz16B/8haj3WXTrTBWtP3zwesMadbTlfW0dgWWDq2rY
QMZLmy1hzzaGpSKryfxOOflTIkrxBur8agfZLRdcxy4A88CxezJgH9EVvnnva0wD
Pl5orc8knwKBgCr7RMDHjpaydLE4kQwdhb92jFPWQEvw1JGM4HlJp+7oUeGirH3Q
XE8XK/AkejrSEFMIs4I0W9VMtItH6huF60D4NKIksBGa98IyLTg4Tsvy+TkS+JLZ
ibRdclz86+okLfMbud4AV1ErNJz59iuDWmFZaELVTonLYJT296Zx5eehAoGABouJ
L63MdO6k0zrcjgQx5CmbPoOamrZ4r4E0DQcdpvJgHanBVjqD9EYVHTMktj5ut3WC
wI16tl60YebYo+bksftaqfYjoBzSHagbxR+GXQEmNz7q3L5zGuXuGq+l1iHvQ7b8
AiHf+/XIm+dKe3YOr+bYE+hUc1n64LAIx0O6zukCgYEAhd1Zb575mTV9IJKrOt0v
pED8QcPNYuabacVrheRyzvAp0SJKqwEsdNu/+Vraz2v5NsHLjKpGuprY+0eH+oI7
aE3JMjytapavUJoHAeGNOzZ9/2Dwybm6qhEK4KyH7sOzc9w0/+g0sJEXyB+ZWCYm
RGcl7ds1oFBlCJbL+AgB5vk=
-----END PRIVATE KEY-----""",
                "client_email": "firebase-adminsdk-fbsvc@softview-cci-utility.iam.gserviceaccount.com",
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
                "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-fbsvc%40softview-cci-utility.iam.gserviceaccount.com"
            }
            cred = credentials.Certificate(firebase_info)
            firebase_admin.initialize_app(cred)
        except Exception as e:
            st.error(f"Database Connection Error: {e}")
            return None
    return firestore.client()

def get_master_data():
    db = init_db()
    records = []
    if db:
        try:
            docs = db.collection("master_data").stream()
            for doc in docs:
                records.append(doc.to_dict())
        except Exception as e:
            st.error(f"Error fetching historical data: {e}")
    return pd.DataFrame(records)

# --- SAFE CASTERS ---
def safe_float(val, default=0.0):
    if pd.isna(val) or val is None:
        return default
    try:
        return float(val)
    except:
        return default

def safe_int(val, default=0):
    if pd.isna(val) or val is None:
        return default
    try:
        return int(float(val))
    except:
        return default

# --- MULTI-ARRAY EXTRACTION FOR EMD & PAYMENTS ---
def extract_emd_and_payments(sheet_df):
    df = sheet_df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    
    emd_list = []
    pay_list = []
    cols = list(df.columns)
    
    c_no_emd = next((i for i, c in enumerate(cols) if 'contract' in c.lower() and i < 3), 0)
    emd_dt_idx = next((i for i, c in enumerate(cols) if 'date' in c.lower() and 'pay' not in c.lower() and i < 4), 1)
    emd_amt_idx = next((i for i, c in enumerate(cols) if ('amt' in c.lower() or 'emd' in c.lower()) and 'pay' not in c.lower() and i < 4), 2)
    
    c_no_pay = next((i for i, c in enumerate(cols) if 'contract' in c.lower() and i >= 3), None)
    if c_no_pay is None:
        c_no_pay = c_no_emd
        
    pay_amt_idx = next((i for i, c in enumerate(cols) if 'payment' in c.lower() and 'date' not in c.lower() and i >= 2), None)
    pay_dt_idx = next((i for i, c in enumerate(cols) if 'date' in c.lower() and 'payment' in c.lower() and i >= 3), None)

    for idx, row in df.iterrows():
        c_emd = str(row.iloc[c_no_emd]).strip().replace('.', '-')
        if c_emd and c_emd != 'nan':
            e_amt = safe_float(row.iloc[emd_amt_idx])
            e_dt = row.iloc[emd_dt_idx]
            if e_amt > 0 and pd.notna(e_dt):
                emd_list.append({'contract': c_emd, 'amt': e_amt, 'date': pd.to_datetime(e_dt)})
                
        if pay_amt_idx is not None and pay_dt_idx is not None:
            c_pay = str(row.iloc[c_no_pay]).strip().replace('.', '-')
            p_amt = safe_float(row.iloc[pay_amt_idx])
            p_dt = row.iloc[pay_dt_idx]
            if p_amt > 0 and pd.notna(p_dt):
                pay_list.append({'contract': c_pay, 'amt': p_amt, 'date': pd.to_datetime(p_dt)})
                
    return pd.DataFrame(emd_list), pd.DataFrame(pay_list)

def calculate_fifo_engine(emd_df, pay_df, master_df, default_rate):
    if emd_df.empty:
        return pd.DataFrame(columns=["Contract No", "Allocated EMD", "Interest", "Days", "Applied Rate (%)"])
        
    interest_records = []
    unique_contracts = emd_df['contract'].unique()
    
    for contract in unique_contracts:
        c_emds = emd_df[emd_df['contract'] == contract].to_dict('records')
        c_pays = []
        if not pay_df.empty and 'contract' in pay_df.columns:
            c_pays = pay_df[pay_df['contract'] == contract].to_dict('records')
            
        current_rate = default_rate
        if not master_df.empty and 'Contract_No' in master_df.columns:
            matched = master_df[master_df['Contract_No'] == str(contract).strip()]
            if not matched.empty:
                current_rate = safe_float(matched.iloc[0].get('EMD_Interest_Rate', default_rate))
                
        p_idx = 0
        for req in c_emds:
            req_amt = req['amt']
            req_date = req['date']
            
            while req_amt > 0 and p_idx < len(c_pays):
                pay = c_pays[p_idx]
                pay_amt = pay['amt']
                pay_date = pay['date']
                
                if pay_amt <= 0:
                    p_idx += 1
                    continue
                    
                allocated = min(req_amt, pay_amt)
                days = (pay_date - req_date).days
                if days < 0 or pd.isna(days): days = 0
                
                interest = (allocated * (current_rate / 100) * days) / 365
                
                interest_records.append({
                    "Contract No": contract,
                    "Allocated EMD": allocated,
                    "Due Date": req_date.strftime('%d-%m-%Y'),
                    "Payment Date": pay_date.strftime('%d-%m-%Y'),
                    "Days": int(days),
                    "Interest": round(interest, 2),
                    "Applied Rate (%)": current_rate
                })
                
                req_amt -= allocated
                c_pays[p_idx]['amt'] -= allocated
                if c_pays[p_idx]['amt'] <= 0:
                    p_idx += 1
                    
    return pd.DataFrame(interest_records) if interest_records else pd.DataFrame(columns=["Contract No", "Allocated EMD", "Interest", "Days", "Applied Rate (%)"])

def get_slab_rate(days, d1, r1, d2, r2, d3, r3):
    if days <= d1: return r1
    elif days <= d2: return r2
    return r3

# --- SIDEBAR MASTER SETUP ---
# =========================================
# PREMIUM ERP STYLE SIDEBAR NAVIGATION
# =========================================

# =========================================
# ULTRA PREMIUM MASTER SIDEBAR
# =========================================

with st.sidebar:

    logo_path = "softview_logo.png"

    if os.path.exists(logo_path):
        try:
            st.image(Image.open(logo_path), use_container_width=True)
        except:
            pass

    st.markdown("""
    <div class="sidebar-header">
        ⚙️ MASTER CONTROL PANEL
    </div>
    """, unsafe_allow_html=True)

    with st.form("master_form_aju_baju_v15"):

        # =====================================
        # BASIC DETAILS
        # =====================================

        with st.expander("📋 BASIC CONTRACT DETAILS", expanded=True):

            p_name = st.text_input(
                "PARTY NAME",
                value=""
            ).strip()

            c_no = st.text_input(
                "CONTRACT NO.",
                value="DEFAULT"
            ).upper().strip().replace('.', '-')

            col1, col2 = st.columns(2)

            with col1:

                c_date = st.date_input(
                    "CONTRACT DATE",
                    value=None,
                    format="DD/MM/YYYY"
                )

                e_dt = st.date_input(
                    "EFFECTIVE DATE",
                    value=None,
                    format="DD/MM/YYYY"
                )

            with col2:

                p_valid = st.date_input(
                    "VALID TILL",
                    value=None,
                    format="DD/MM/YYYY"
                )

                emd_slab = st.number_input(
                    "EMD SLAB (%)",
                    value=0.0
                )

        # =====================================
        # CD SETTINGS
        # =====================================

        with st.expander("💰 CASH DISCOUNT SETTINGS", expanded=False):

            col1, col2 = st.columns(2)

            with col1:

                cd_d1 = st.number_input(
                    "CD DAYS",
                    value=0
                )

            with col2:

                cd_r1 = st.number_input(
                    "CD RATE (%)",
                    value=0.0
                )

            tax_rate = st.number_input(
                "GST / TAX RATE (%)",
                value=18.0
            )

        # =====================================
        # LL SETTINGS
        # =====================================

        with st.expander("📉 LATE LIFTING SETTINGS", expanded=False):

            col1, col2 = st.columns(2)

            with col1:

                ll_p1 = st.number_input("LL Days 1", value=0)
                ll_p2 = st.number_input("LL Days 2", value=0)
                ll_p3 = st.number_input("LL Days 3", value=0)

            with col2:

                ll_r1 = st.number_input("LL Rate 1", value=0.0)
                ll_r2 = st.number_input("LL Rate 2", value=0.0)
                ll_r3 = st.number_input("LL Rate 3", value=0.0)

        # =====================================
        # CC SETTINGS
        # =====================================

        with st.expander("🚛 CARRYING COST SETTINGS", expanded=False):

            col1, col2 = st.columns(2)

            with col1:

                cc_d1 = st.number_input("CC Days 1", value=0)
                cc_d2 = st.number_input("CC Days 2", value=0)
                cc_d3 = st.number_input("CC Days 3", value=0)

            with col2:

                cc_r1 = st.number_input("CC Rate 1", value=0.0)
                cc_r2 = st.number_input("CC Rate 2", value=0.0)
                cc_r3 = st.number_input("CC Rate 3", value=0.0)

        # =====================================
        # EMD SETTINGS
        # =====================================

        with st.expander("🏦 EMD INTEREST SETTINGS", expanded=False):

            emd_int_rate = st.number_input(
                "EMD INTEREST RATE (%)",
                value=0.0
            )

        # =====================================
        # SAVE BUTTON
        # =====================================

        submitted = st.form_submit_button(
            "🏆 SAVE MASTER CONFIGURATION"
        )

        if submitted and c_no:

            db = init_db()

            if db:

                doc_id = (
                    f"CFG_{c_no}"
                    if c_no == "DEFAULT"
                    else f"{p_name}_{c_no}"
                )

                data = {

                    "Party_Name":
                        p_name if p_name else "DEFAULT_GLOBAL",

                    "Contract_No": c_no,

                    "Project_Valid_Till":
                        p_valid.strftime('%d-%m-%Y')
                        if p_valid else "",

                    "Contract_Date":
                        c_date.strftime('%d-%m-%Y')
                        if c_date else "",

                    "Effective_Date":
                        e_dt.strftime('%d-%m-%Y')
                        if e_dt else "",

                    "EMD_Slab_Pct": emd_slab,
                    "EMD_Interest_Rate": emd_int_rate,
                    "Tax_Rate": tax_rate,

                    "CD_Days_1": cd_d1,
                    "CD_Rate_1": cd_r1,

                    "LL_Days_1": ll_p1,
                    "LL_Rate_1": ll_r1,

                    "LL_Days_2": ll_p2,
                    "LL_Rate_2": ll_r2,

                    "LL_Days_3": ll_p3,
                    "LL_Rate_3": ll_r3,

                    "CC_Days_1": cc_d1,
                    "CC_Rate_1": cc_r1,

                    "CC_Days_2": cc_d2,
                    "CC_Rate_2": cc_r2,

                    "CC_Days_3": cc_d3,
                    "CC_Rate_3": cc_r3
                }

                db.collection("master_data").document(doc_id).set(data)

                st.success(
                    f"✅ Configuration Saved : {c_no}"
                )

                st.rerun()


# --- MAIN ENGINE CONTENT ---
st.title("⚜️ CCI CALCULATION WORKING UTILITY ⚜️")
master_hist_df = get_master_data()

if not master_hist_df.empty:
    if 'Contract_No' in master_hist_df.columns:
        master_hist_df['Contract_No'] = master_hist_df['Contract_No'].astype(str).str.strip().str.replace('.', '-', regex=False)
    st.dataframe(master_hist_df, use_container_width=True)

st.markdown("---")
st.markdown("### 📁 BATCH EXTRACTION FILES UPLOADER")
uploaded_file = st.file_uploader("Upload 'cci_working_upload_sheet.xlsx':", type=["xlsx"])

if uploaded_file is not None:
    st.info("Excel sheets successfully uploaded in background buffer.")
    
    run_calc = st.button("🔄 RUN CALCULATION ENGINE")
    
    if run_calc:
        try:
            pur_df = pd.read_excel(uploaded_file, sheet_name="PUR CONT DETAILS")
            emd_payment_sheet = pd.read_excel(uploaded_file, sheet_name="EMD PAYMENT DETAILS ")
            grn_df = pd.read_excel(uploaded_file, sheet_name="GRN BOOKING ")
            
            pur_df.columns = [str(c).strip() for c in pur_df.columns]
            grn_df.columns = [str(c).strip() for c in grn_df.columns]
            
            pur_contract_key = 'Contract No.'
            pur_eff_date_key = 'EFFECTIVE DATE'
            grn_contract_key = 'contract no'
            grn_bill_date_key = 'Party Bill Date'
            grn_qty_key = 'Accepted Quantity(AUM)'  
            grn_mat_amt_key = 'Material Amount'
            grn_id_key = 'GRN'

            clean_emd_df, clean_pay_df = extract_emd_and_payments(emd_payment_sheet)
            
            base_emd_rate = 0.0
            def_row = master_hist_df[master_hist_df['Contract_No'] == "DEFAULT"] if not master_hist_df.empty else pd.DataFrame()
            if not def_row.empty:
                base_emd_rate = safe_float(def_row.iloc[0].get('EMD_Interest_Rate', 0.0))
            
            fifo_res = calculate_fifo_engine(clean_emd_df, clean_pay_df, master_hist_df, base_emd_rate)
            
            pur_clean = pur_df[[pur_contract_key, pur_eff_date_key]].dropna().copy()
            pur_clean['clean_cno'] = pur_clean[pur_contract_key].astype(str).str.strip().str.replace('.', '-', regex=False)

            ll_charges_list = []
            cc_charges_list = []
            cd_charges_list = []
            emd_int_mapping_list = []
            total_days_list = []
            eff_dates_list = []
            mat_amt_sum = 0.0
            
            contract_emd_totals = {}
            if not fifo_res.empty:
                contract_emd_totals = fifo_res.groupby("Contract No")["Interest"].sum().to_dict()

            for idx, row in grn_df.iterrows():
                c_id = str(row[grn_contract_key]).strip().replace('.', '-')
                qty = safe_float(row.get(grn_qty_key, 0.0))
                mat_amt = safe_float(row.get(grn_mat_amt_key, 0.0))
                mat_amt_sum += mat_amt
                
                pur_match = pur_clean[pur_clean['clean_cno'] == c_id]
                eff_dt = pur_match.iloc[0][pur_eff_date_key] if not pur_match.empty else None
                bill_dt = row.get(grn_bill_date_key, None)
                
                try:
                    b_dt = pd.to_datetime(bill_dt)
                    e_dt = pd.to_datetime(eff_dt)
                    days = (b_dt - e_dt).days
                    if pd.isna(days) or days < 0: days = 0
                except:
                    days = 0
                    e_dt = pd.NaT
                    
                total_days_list.append(days)
                eff_dates_list.append(e_dt)
                
                rule = master_hist_df[master_hist_df['Contract_No'] == c_id] if not master_hist_df.empty else pd.DataFrame()
                if rule.empty: rule = def_row
                    
                if not rule.empty:
                    r = rule.iloc[0]
                    ll_days_1 = safe_int(r.get('LL_Days_1', 0)); ll_rate_1 = safe_float(r.get('LL_Rate_1', 0.0))
                    ll_days_2 = safe_int(r.get('LL_Days_2', 0)); ll_rate_2 = safe_float(r.get('LL_Rate_2', 0.0))
                    ll_days_3 = safe_int(r.get('LL_Days_3', 0)); ll_rate_3 = safe_float(r.get('LL_Rate_3', 0.0))
                    
                    cc_days_1 = safe_int(r.get('CC_Days_1', 0)); cc_rate_1 = safe_float(r.get('CC_Rate_1', 0.0))
                    cc_days_2 = safe_int(r.get('CC_Days_2', 0)); cc_rate_2 = safe_float(r.get('CC_Rate_2', 0.0))
                    cc_days_3 = safe_int(r.get('CC_Days_3', 0)); cc_rate_3 = safe_float(r.get('CC_Rate_3', 0.0))
                    
                    g_cd = safe_float(r.get('CD_Rate_1', 0.0))

                    l_days = max(0, days - ll_days_1)
                    l_rate = get_slab_rate(days, ll_days_1, ll_rate_1, ll_days_2, ll_rate_2, ll_days_3, ll_rate_3)
                    ll_charges_list.append(qty * l_days * l_rate)
                    
                    c_rate = get_slab_rate(days, cc_days_1, cc_rate_1, cc_days_2, cc_rate_2, cc_days_3, cc_rate_3)
                    cc_charges_list.append(qty * days * c_rate)
                    
                    cd_charges_list.append(mat_amt * (g_cd / 100))
                else:
                    ll_charges_list.append(0.0)
                    cc_charges_list.append(0.0)
                    cd_charges_list.append(0.0)
                
                emd_int_mapping_list.append(contract_emd_totals.get(c_id, 0.0) / len(grn_df[grn_df[grn_contract_key] == row[grn_contract_key]]) if c_id in contract_emd_totals else 0.0)
                    
            grn_df['Effective_Date'] = eff_dates_list
            grn_df['Total_Days'] = total_days_list
            grn_df['LL_Charges'] = ll_charges_list
            grn_df['CC_Charges'] = cc_charges_list
            grn_df['CD_Amount'] = cd_charges_list
            grn_df['EMD_Interest_Allocated'] = emd_int_mapping_list
            
            total_bales = int(safe_float(grn_df[grn_qty_key].sum()))
            total_mat_amt = mat_amt_sum
            t_ll = sum(ll_charges_list)
            t_cc = sum(cc_charges_list)
            t_cd = sum(cd_charges_list)
            total_emd_int = fifo_res['Interest'].sum() if not fifo_res.empty else 0.0
            
            g_tax = safe_float(def_row.iloc[0].get('Tax_Rate', 18.0)) if not def_row.empty else 18.0
            calculated_gst = (t_ll + t_cc) * (g_tax / 100)
            
            total_payments = clean_pay_df['amt'].sum() if not clean_pay_df.empty else 0.0
            net_payable = (total_mat_amt + t_ll + t_cc + calculated_gst) - (t_cd + total_emd_int + total_payments)

            # --- RENDER BLOCK OUTPUTS ---
            st.markdown("## 🏆 FINAL CALCULATION AUDIT REPORT")
            tab1, tab2, tab3 = st.tabs(["📊 1. EXECUTIVE SUMMARY", "📋 2. DETAILED LEDGER MATRIX", "📜 3. TEMPLATE INVOICE REPORT"])
            
            with tab1:
                c_m1, c_m2, c_m3 = st.columns(3)
                with c_m1: st.markdown(f'<div class="luxury-card"><div class="luxury-metric-title">Total Bales Lifted</div><div class="luxury-metric-value">{total_bales:,} Bales</div></div>', unsafe_allow_html=True)
                with c_m2: st.markdown(f'<div class="luxury-card"><div class="luxury-metric-title">Gross Material Value</div><div class="luxury-metric-value">₹ {total_mat_amt:,.2f}</div></div>', unsafe_allow_html=True)
                with c_m3: st.markdown(f'<div class="luxury-card"><div class="luxury-metric-title">Net Payable / Receivable</div><div class="luxury-metric-value">₹ {net_payable:,.2f}</div></div>', unsafe_allow_html=True)
                
                summary_data = {
                    "Particulars Label": ["Material Amount Base", "Late Lifting (LL) Penalties", "Carrying Cost (CC) Charges", "Accrued GST On Charges", "Cash Discount (CD)", "FIFO EMD Interest Rebate", "Total Payments Received"],
                    "Total Value (₹)": [total_mat_amt, t_ll, t_cc, calculated_gst, t_cd, total_emd_int, total_payments]
                }
                st.dataframe(pd.DataFrame(summary_data), use_container_width=True)

            with tab2:
                final_grn = grn_df[[grn_contract_key, grn_id_key, 'Effective_Date', grn_bill_date_key, grn_qty_key, 'Total_Days', 'LL_Charges', 'CC_Charges', 'CD_Amount', 'EMD_Interest_Allocated']].copy()
                final_grn['Effective_Date'] = pd.to_datetime(final_grn['Effective_Date']).dt.strftime('%d-%m-%Y')
                final_grn[grn_bill_date_key] = pd.to_datetime(final_grn[grn_bill_date_key]).dt.strftime('%d-%m-%Y')
                final_grn.rename(columns={'CD_Amount': 'CD Amount (₹)', 'EMD_Interest_Allocated': 'EMD Interest (₹)'}, inplace=True)
                st.dataframe(final_grn, use_container_width=True)

            with tab3:
                col_exp1, col_exp2 = st.columns(2)
                
                # --- HIGHLY DESIGNED COLOURFUL EXCEL BUILDER ---
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df_invoice_export = pd.DataFrame({
                        "SR. NO.": [1, 2, 3, 4, 5, 6, 7, 8],
                        "PARTICULARS": [
                            f"Gross Material Amount (Total Bales: {total_bales:,})",
                            "Carrying Cost (CC) Charges",
                            "Late Lifting (LL) Penalties",
                            f"Add: GST on CC & LL Charges ({g_tax}%)",
                            "GROSS TOTAL (A)",
                            "Less: Cash Discount (CD) Received",
                            "Less: EMD Interest Rebate (FIFO Method)",
                            "Less: Total EMD / Payments Realized"
                        ],
                        "RATE / %": ["Base Price", "Slab Wise", "Slab Wise", f"{g_tax}%", "-", "Config Rate", "As per Master", "Actual Paid"],
                        "AMOUNT (₹)": [
                            total_mat_amt, t_cc, t_ll, calculated_gst, 
                            (total_mat_amt + t_cc + t_ll + calculated_gst),
                            -t_cd, -total_emd_int, -total_payments
                        ]
                    })
                    df_invoice_export.to_excel(writer, sheet_name="Settlement Statement", index=False)
                    
                    workbook = writer.book
                    worksheet = writer.sheets["Settlement Statement"]
                    
                    purple_header_fill = PatternFill(start_color="8A2BE2", end_color="8A2BE2", fill_type="solid")
                    gross_total_fill = PatternFill(start_color="F1EAFA", end_color="F1EAFA", fill_type="solid") 
                    net_payable_fill = PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid") 
                    
                    white_bold_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                    bold_black_font = Font(name="Segoe UI", size=11, bold=True)
                    regular_font = Font(name="Segoe UI", size=11)
                    final_net_font = Font(name="Segoe UI", size=12, bold=True, color="8A2BE2")
                    
                    thin_border_side = Side(border_style="thin", color="D3D3D3")
                    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
                    
                    for col_idx in range(1, 5):
                        cell = worksheet.cell(row=1, column=col_idx)
                        cell.fill = purple_header_fill
                        cell.font = white_bold_font
                        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
                    
                    for row_idx in range(2, 10):
                        for col_idx in range(1, 5):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.font = regular_font
                            cell.border = grid_border
                            if col_idx == 1:
                                cell.alignment = Alignment(horizontal="center")
                            elif col_idx == 4:
                                cell.number_format = '₹ #,##0.00'
                                cell.alignment = Alignment(horizontal="right")
                                
                        if row_idx == 6:
                            for col_idx in range(1, 5):
                                worksheet.cell(row=row_idx, column=col_idx).fill = gross_total_fill
                                worksheet.cell(row=row_idx, column=col_idx).font = bold_black_font
                                
                    for col_idx in range(1, 5):
                        worksheet.cell(row=10, column=col_idx).fill = net_payable_fill
                        worksheet.cell(row=10, column=col_idx).font = final_net_font
                        
                    for col in worksheet.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 15)
                        
                excel_buffer.seek(0)
                
                # --- NATIVE PDF STREAM GENERATION ---
                pdf = FPDF()
                pdf.add_page()
                pdf.set_text_color(138, 43, 226)
                pdf.set_font("Arial", "B", 16)
                pdf.cell(190, 10, "THE COTTON CORPORATION OF INDIA LTD.", ln=True, align="C")
                pdf.set_text_color(85, 85, 85)
                pdf.set_font("Arial", "B", 12)
                pdf.cell(190, 8, "FINAL SETTLEMENT STATEMENT", ln=True, align="C")
                pdf.ln(10)
                
                pdf.set_fill_color(138, 43, 226)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font("Arial", "B", 10)
                pdf.cell(15, 10, "SR.", border=1, fill=True, align="C")
                pdf.cell(100, 10, "PARTICULARS", border=1, fill=True)
                pdf.cell(35, 10, "RATE / %", border=1, fill=True, align="C")
                pdf.cell(40, 10, "AMOUNT (Rs.)", border=1, fill=True, align="R")
                pdf.ln()
                
                pdf.set_text_color(0, 0, 0)
                pdf.set_font("Arial", "", 10)
                
                rows_pdf = [
                    ("1", f"Gross Material Value ({total_bales:,} Bales)", "Base Price", f"{total_mat_amt:,.2f}"),
                    ("2", "Carrying Cost (CC) Charges", "Slab Wise", f"{t_cc:,.2f}"),
                    ("3", "Late Lifting (LL) Penalties", "Slab Wise", f"{t_ll:,.2f}"),
                    ("4", f"Add: GST on CC & LL ({g_tax}%)", f"{g_tax}%", f"{calculated_gst:,.2f}"),
                    ("", "GROSS TOTAL (A)", "-", f"{(total_mat_amt + t_cc + t_ll + calculated_gst):,.2f}"),
                    ("5", "Less: Cash Discount (CD) Received", "Config Rate", f"-{t_cd:,.2f}"),
                    ("6", "Less: EMD Interest Rebate (FIFO)", "As per Master", f"-{total_emd_int:,.2f}"),
                    ("7", "Less: Total Payments Realized", "Actual Paid", f"-{total_payments:,.2f}"),
                ]
                
                for sr, part, rate, amt in rows_pdf:
                    if part == "GROSS TOTAL (A)":
                        pdf.set_fill_color(241, 234, 250)
                        pdf.set_font("Arial", "B", 10)
                        pdf.cell(15, 8, sr, border=1, fill=True, align="C")
                        pdf.cell(100, 8, part, border=1, fill=True)
                        pdf.cell(35, 8, rate, border=1, fill=True, align="C")
                        pdf.cell(40, 8, amt, border=1, fill=True, align="R")
                        pdf.set_font("Arial", "", 10)
                    else:
                        pdf.cell(15, 8, sr, border=1, align="C")
                        pdf.cell(100, 8, part, border=1)
                        pdf.cell(35, 8, rate, border=1, align="C")
                        pdf.cell(40, 8, amt, border=1, align="R")
                    pdf.ln()
                    
                pdf.set_fill_color(255, 240, 245)
                pdf.set_text_color(138, 43, 226)
                pdf.set_font("Arial", "B", 11)
                pdf.cell(15, 10, "", border=1, fill=True)
                pdf.cell(100, 10, "NET PAYABLE / (RECEIVABLE) DUE", border=1, fill=True)
                pdf.cell(35, 10, "-", border=1, fill=True, align="C")
                pdf.cell(40, 10, f"Rs. {net_payable:,.2f}", border=1, fill=True, align="R")
                
                pdf_bytes = pdf.output(dest='B')
                pdf_buffer = io.BytesIO(pdf_bytes)

                with col_exp1:
                    st.download_button(
                        label="📥 EXCEL REPORT GENERATE",
                        data=excel_buffer,
                        file_name=f"CCI_Final_Settlement_Report_{datetime.now().strftime('%d%m%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with col_exp2:
                    st.download_button(
                        label="📄 PDF REPORT GENERATE",
                        data=pdf_buffer,
                        file_name=f"CCI_Final_Settlement_Statement_{datetime.now().strftime('%d%m%Y')}.pdf",
                        mime="application/pdf"
                    )
                    
                st.markdown("---")
                
                invoice_html = f"""
                <div style="background-color: #FFFFFF; padding: 25px; border: 1px solid #8A2BE2; border-radius: 8px;">
                    <div style="text-align: center; font-size: 20px; font-weight: bold; color: #8A2BE2; margin-bottom: 20px;">
                        THE COTTON CORPORATION OF INDIA LTD. <br>
                        <span style="font-size: 14px; color: #555555;">FINAL SETTLEMENT STATEMENT</span>
                    </div>
                    <table class="invoice-table">
                        <thead>
                            <tr>
                                <th>SR. NO.</th>
                                <th>PARTICULARS</th>
                                <th style="text-align: right;">RATE / %</th>
                                <th style="text-align: right;">AMOUNT (₹)</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td>1</td>
                                <td>Gross Material Amount (Total Bales: {total_bales:,})</td>
                                <td style="text-align: right;">Base Price</td>
                                <td style="text-align: right;">₹ {total_mat_amt:,.2f}</td>
                            </tr>
                            <tr>
                                <td>2</td>
                                <td>Carrying Cost (CC) Charges</td>
                                <td style="text-align: right;">Slab Wise</td>
                                <td style="text-align: right;">₹ {t_cc:,.2f}</td>
                            </tr>
                            <tr>
                                <td>3</td>
                                <td>Late Lifting (LL) Penalties</td>
                                <td style="text-align: right;">Slab Wise</td>
                                <td style="text-align: right;">₹ {t_ll:,.2f}</td>
                            </tr>
                            <tr>
                                <td>4</td>
                                <td>Add: GST on CC & LL Charges</td>
                                <td style="text-align: right;">{g_tax}%</td>
                                <td style="text-align: right;">₹ {calculated_gst:,.2f}</td>
                            </tr>
                            <tr style="background-color: #F1EAFA; font-weight: bold;">
                                <td></td>
                                <td>GROSS TOTAL (A)</td>
                                <td style="text-align: right;">-</td>
                                <td style="text-align: right;">₹ {(total_mat_amt + t_cc + t_ll + calculated_gst):,.2f}</td>
                            </tr>
                            <tr>
                                <td>5</td>
                                <td>Less: Cash Discount (CD) Received</td>
                                <td style="text-align: right;">Config Rate</td>
                                <td style="text-align: right; color: red;">- ₹ {t_cd:,.2f}</td>
                            </tr>
                            <tr>
                                <td>6</td>
                                <td>Less: EMD Interest Rebate (FIFO Method)</td>
                                <td style="text-align: right;">As per Master</td>
                                <td style="text-align: right; color: red;">- ₹ {total_emd_int:,.2f}</td>
                            </tr>
                            <tr>
                                <td>7</td>
                                <td>Less: Total EMD / Payments Realized</td>
                                <td style="text-align: right;">Actual Paid</td>
                                <td style="text-align: right; color: red;">- ₹ {total_payments:,.2f}</td>
                            </tr>
                            <tr class="invoice-total-row" style="background-color: #FFF0F5;">
                                <td></td>
                                <td style="color: #8A2BE2; font-size: 16px;">NET PAYABLE / (RECEIVABLE) DUE</td>
                                <td style="text-align: right;">-</td>
                                <td style="text-align: right; font-size: 16px; color: #8A2BE2;">₹ {net_payable:,.2f}</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                """
                st.markdown(invoice_html, unsafe_allow_html=True)
                
        except Exception as e:
            st.error(f"Calculation Error: {e}")
